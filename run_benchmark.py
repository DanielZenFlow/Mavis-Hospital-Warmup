#!/usr/bin/env python3
"""
Benchmark script for the MAvis Hospital search client.

Usage:
    python run_benchmark.py --xlsx warmup_Benchmark.xlsx          # Run all tasks defined in Excel
    python run_benchmark.py --xlsx warmup_Benchmark.xlsx -t 300   # Custom timeout
    python run_benchmark.py --xlsx warmup_Benchmark.xlsx --compile # Compile first
    python run_benchmark.py -s bfs -f MAPF                        # Legacy: manual strategy + filter
"""

import argparse
import glob
import os
import subprocess
import sys
import time

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Configuration ────────────────────────────────────────────────────────────
LEVEL_DIR = "levels"
OUTPUT_FILE = "benchmark_results.md"
SERVER_JAR = "server.jar"
DEFAULT_TIMEOUT = 180          # seconds per level
JAVA_XMX = "4g"
DEFAULT_STRATEGY = "bfs"
VALID_STRATEGIES = ["bfs", "dfs", "astar", "wastar", "greedy"]


# ── Helpers ──────────────────────────────────────────────────────────────────

def compile_client():
    """Compile the search client with javac."""
    print("Compiling searchclient/*.java ...")
    result = subprocess.run(
        ["javac", "searchclient/*.java"],
        shell=True,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print("ERROR: Compilation failed!")
        print(result.stderr)
        sys.exit(1)
    print("Compilation successful.\n")


def parse_output(output: str):
    """
    Parse combined stdout+stderr from the server / client for key metrics.

    Looks for:
        Found solution of length <N>.          (client stderr)
        #Expanded: … #Frontier: … Time: …     (client stderr)
        Memory used: …                         (client stderr)
    """
    solved = False
    solution_length = "-"
    solve_time = "-"
    memory = "-"
    expanded = "-"
    frontier_size = "-"
    generated = "-"

    for line in output.splitlines():
        if "Found solution of length" in line:
            solved = True
            try:
                solution_length = line.split("length")[1].strip().strip(".,").replace(",", "")
            except Exception:
                pass

        if "#Expanded:" in line:
            try:
                # Use regex to handle numbers with commas (e.g. 858,035)
                import re
                m_exp = re.search(r"#Expanded:\s*([\d,]+)", line)
                m_frt = re.search(r"#Frontier:\s*([\d,]+)", line)
                m_gen = re.search(r"#Generated:\s*([\d,]+)", line)
                m_time = re.search(r"Time:\s*([\d.]+)\s*s", line)
                if m_exp:
                    expanded = m_exp.group(1).replace(",", "")
                if m_frt:
                    frontier_size = m_frt.group(1).replace(",", "")
                if m_gen:
                    generated = m_gen.group(1).replace(",", "")
                if m_time:
                    solve_time = m_time.group(1)
            except Exception:
                pass

        if "Memory used:" in line:
            try:
                memory = line.split("Memory used:")[1].strip()
            except Exception:
                pass

        if "Unable to solve level" in line:
            solved = False

    return {
        "solved": solved,
        "solution_length": solution_length,
        "time": solve_time,
        "memory": memory,
        "expanded": expanded,
        "generated": generated,
    }


def run_level(level_path: str, strategy: str, timeout: int):
    """Run a single level through the server and return parsed metrics."""
    client_cmd = f"java -Xmx{JAVA_XMX} searchclient.SearchClient -{strategy}"
    cmd = [
        "java", "-jar", SERVER_JAR,
        "-l", level_path,
        "-c", client_cmd,
        "-t", str(timeout),
    ]

    level_name = os.path.basename(level_path)
    print(f"  {level_name:<40s}", end="", flush=True)

    wall_start = time.time()
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout + 10,      # extra grace period
        )
        wall_time = time.time() - wall_start
        full_output = (result.stdout or "") + "\n" + (result.stderr or "")

        metrics = parse_output(full_output)

        if result.returncode != 0 and not metrics["solved"]:
            metrics["status"] = "❌ Error"
            print(f"  Error (exit {result.returncode})")
        elif metrics["solved"]:
            metrics["status"] = "✅ Solved"
            print(f"  Solved  len={metrics['solution_length']:>6s}  t={metrics['time']:>8s}s")
        else:
            metrics["status"] = "❌ No solution"
            print(f"  No solution  t={metrics['time']:>8s}s")

        metrics["wall_time"] = f"{wall_time:.1f}"
        return metrics

    except subprocess.TimeoutExpired:
        print(f"  ⏱️  Timeout (>{timeout}s)")
        return {
            "status": "⏱️ Timeout",
            "solved": False,
            "solution_length": "-",
            "time": f">{timeout}",
            "wall_time": f">{timeout}",
            "memory": "-",
            "expanded": "-",
            "generated": "-",
        }
    except Exception as e:
        print(f"  Exception: {e}")
        return {
            "status": "❌ Exception",
            "solved": False,
            "solution_length": "-",
            "time": "-",
            "wall_time": "-",
            "memory": "-",
            "expanded": "-",
            "generated": "-",
        }


# ── Excel helpers ────────────────────────────────────────────────────────────

def read_xlsx_tasks(xlsx_path: str):
    """Read (level, strategy) task pairs from the Excel file.

    Returns a list of dicts: [{"level": "MAPF00", "strategy": "bfs", "sheet": ..., "row": ...}, ...]
    """
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        sys.exit(1)
    if not os.path.exists(xlsx_path):
        print(f"ERROR: {xlsx_path} not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    tasks = []
    for ws in wb.worksheets:
        for row_idx in range(2, ws.max_row + 1):
            cell_level = ws.cell(row=row_idx, column=1).value
            cell_strategy = ws.cell(row=row_idx, column=2).value
            if cell_level and cell_strategy:
                tasks.append({
                    "level": cell_level.strip(),
                    "strategy": cell_strategy.strip().lower(),
                    "sheet": ws.title,
                    "row": row_idx,
                })
    wb.close()
    return tasks


def write_xlsx_results(xlsx_path: str, task_results: list):
    """Write benchmark results back to the Excel file.

    task_results: list of (task_dict, metrics_dict).
    """
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.load_workbook(xlsx_path)
    updated = 0

    for task, metrics in task_results:
        if metrics is None:
            continue
        ws = wb[task["sheet"]]
        row = task["row"]
        ws.cell(row=row, column=3).value = _to_number(metrics.get("generated", "-"))
        ws.cell(row=row, column=4).value = _to_number(metrics.get("time", "-"))
        ws.cell(row=row, column=5).value = _to_number(metrics.get("solution_length", "-"))
        updated += 1

    wb.save(xlsx_path)
    print(f"\nExcel: updated {updated} row(s) in {xlsx_path}")


def _to_number(val):
    """Convert string metric to int/float if possible, else return as-is."""
    if val is None or val == "-":
        return None
    try:
        if "." in str(val):
            return float(val)
        return int(val)
    except (ValueError, TypeError):
        return val


def find_level_file(level_name: str):
    """Find the .lvl file for a given level name. Returns path or None."""
    path = os.path.join(LEVEL_DIR, f"{level_name}.lvl")
    if os.path.exists(path):
        return path
    return None


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="MAvis Hospital benchmark runner")
    parser.add_argument("--xlsx", default=None,
                        help="Path to warmup_Benchmark.xlsx — run exactly the tasks defined in it")
    parser.add_argument("-s", "--strategy", default=DEFAULT_STRATEGY,
                        choices=VALID_STRATEGIES,
                        help="Search strategy, used only in legacy mode without --xlsx (default: bfs)")
    parser.add_argument("-t", "--timeout", type=int, default=DEFAULT_TIMEOUT,
                        help="Timeout per level in seconds (default: 180)")
    parser.add_argument("-f", "--filter", default=None,
                        help="Only run levels whose filename starts with this prefix (legacy mode)")
    parser.add_argument("-o", "--output", default=OUTPUT_FILE,
                        help="Output markdown file (default: benchmark_results.md)")
    parser.add_argument("--compile", action="store_true",
                        help="Compile searchclient before running")
    args = parser.parse_args()

    # ── Prerequisites ────────────────────────────────────────────────────
    if not os.path.exists(SERVER_JAR):
        print(f"ERROR: {SERVER_JAR} not found in {os.getcwd()}")
        sys.exit(1)

    if args.compile:
        compile_client()

    class_files = glob.glob("searchclient/*.class")
    if not class_files:
        print("ERROR: No .class files in searchclient/. Compile first:")
        print("  javac searchclient/*.java")
        print("  or run with --compile flag")
        sys.exit(1)

    # ── Excel-driven mode ────────────────────────────────────────────────
    if args.xlsx:
        run_xlsx_mode(args)
    else:
        run_legacy_mode(args)


def run_xlsx_mode(args):
    """Run benchmark tasks as defined in the Excel file."""
    tasks = read_xlsx_tasks(args.xlsx)

    if not tasks:
        print("ERROR: No tasks found in Excel file.")
        sys.exit(1)

    # Summarize
    strategies = sorted(set(t["strategy"] for t in tasks))
    levels = sorted(set(t["level"] for t in tasks))
    print(f"{'=' * 60}")
    print(f"  MAvis Hospital Benchmark (Excel-driven)")
    print(f"  Excel    : {args.xlsx}")
    print(f"  Tasks    : {len(tasks)} ({len(levels)} levels × strategies: {', '.join(strategies)})")
    print(f"  Timeout  : {args.timeout}s per level")
    print(f"  CWD      : {os.getcwd()}")
    print(f"{'=' * 60}\n")

    # Run all tasks
    task_results = []   # list of (task, metrics)
    solved_count = 0
    skip_count = 0
    timeout_count = 0
    error_count = 0

    for task in tasks:
        level_name = task["level"]
        strategy = task["strategy"]
        lvl_path = find_level_file(level_name)

        print(f"[{strategy.upper():>5s}] ", end="")

        if lvl_path is None:
            print(f"  {level_name:<40s}  ⚠️  Level file not found, skipped")
            task_results.append((task, None))
            skip_count += 1
            continue

        metrics = run_level(lvl_path, strategy, args.timeout)
        task_results.append((task, metrics))

        if metrics["solved"]:
            solved_count += 1
        elif "Timeout" in metrics["status"]:
            timeout_count += 1
        else:
            error_count += 1

    # ── Write results back to Excel ──────────────────────────────────────
    write_xlsx_results(args.xlsx, task_results)

    # ── Generate markdown report ─────────────────────────────────────────
    effective_tasks = [tr for tr in task_results if tr[1] is not None]
    lines = []
    lines.append(f"# Benchmark Results\n")
    lines.append(f"**Date:** {time.strftime('%Y-%m-%d %H:%M:%S')}  ")
    lines.append(f"**Source:** `{args.xlsx}` | **Timeout:** {args.timeout}s | **Java Xmx:** {JAVA_XMX}  ")
    lines.append(f"**Score:** {solved_count}/{len(effective_tasks)} solved")
    if skip_count:
        lines.append(f" | {skip_count} skipped (no .lvl)")
    if timeout_count:
        lines.append(f" | {timeout_count} timeout")
    if error_count:
        lines.append(f" | {error_count} error/unsolved")
    lines.append("\n")
    lines.append("| Level | Strategy | Status | Solution Length | Time (s) | Expanded | Generated | Memory |")
    lines.append("|-------|----------|--------|-----------------|----------|----------|-----------|--------|")

    for task, m in task_results:
        if m is None:
            lines.append(
                f"| `{task['level']}` "
                f"| {task['strategy'].upper()} "
                f"| ⚠️ Skipped "
                f"| - | - | - | - | - |"
            )
        else:
            lines.append(
                f"| `{task['level']}` "
                f"| {task['strategy'].upper()} "
                f"| {m['status']} "
                f"| {m['solution_length']} "
                f"| {m['time']} "
                f"| {m['expanded']} "
                f"| {m['generated']} "
                f"| {m['memory']} |"
            )

    lines.append("")
    lines.append("### Summary")
    lines.append(f"- **Total tasks**: {len(tasks)}")
    lines.append(f"- **Solved**: {solved_count} ✅")
    lines.append(f"- **Skipped** (missing .lvl): {skip_count} ⚠️")
    lines.append(f"- **Timeout**: {timeout_count} ⏱️")
    lines.append(f"- **Error / Unsolved**: {error_count} ❌")

    report = "\n".join(lines) + "\n"

    try:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"\n{'=' * 60}")
        print(f"  Results saved to {args.output}")
        print(f"  Score: {solved_count}/{len(effective_tasks)} solved"
              f"  ({skip_count} skipped)")
        print(f"{'=' * 60}")
    except Exception as e:
        print(f"ERROR writing results file: {e}")
        sys.exit(1)


def run_legacy_mode(args):
    """Original mode: run all levels matching -f filter with a single -s strategy."""
    print(f"{'=' * 60}")
    print(f"  MAvis Hospital Benchmark (legacy mode)")
    print(f"  Strategy : {args.strategy}")
    print(f"  Timeout  : {args.timeout}s per level")
    print(f"  Filter   : {args.filter or 'all levels'}")
    print(f"  CWD      : {os.getcwd()}")
    print(f"  Python   : {sys.version.split()[0]}")
    print(f"{'=' * 60}\n")

    # Discover levels
    pattern = os.path.join(LEVEL_DIR, "*.lvl")
    levels = sorted(glob.glob(pattern))

    if args.filter:
        levels = [l for l in levels if os.path.basename(l).startswith(args.filter)]

    if not levels:
        print(f"No levels found matching pattern in {LEVEL_DIR}/")
        sys.exit(1)

    print(f"Found {len(levels)} level(s).\n")

    # Run benchmark
    results = []
    solved_count = 0
    timeout_count = 0
    error_count = 0

    for lvl in levels:
        level_name = os.path.basename(lvl).replace(".lvl", "")
        metrics = run_level(lvl, args.strategy, args.timeout)
        results.append((level_name, metrics))
        if metrics["solved"]:
            solved_count += 1
        elif "Timeout" in metrics["status"]:
            timeout_count += 1
        else:
            error_count += 1

    # Generate markdown report
    lines = []
    lines.append(f"# Benchmark Results\n")
    lines.append(f"**Date:** {time.strftime('%Y-%m-%d %H:%M:%S')}  ")
    lines.append(f"**Strategy:** `{args.strategy}` | **Timeout:** {args.timeout}s | **Java Xmx:** {JAVA_XMX}  ")
    lines.append(f"**Score:** {solved_count}/{len(levels)} solved")
    if timeout_count:
        lines.append(f" | {timeout_count} timeout")
    if error_count:
        lines.append(f" | {error_count} error/unsolved")
    lines.append("\n")
    lines.append("| Level | Status | Solution Length | Time (s) | Expanded | Generated | Memory |")
    lines.append("|-------|--------|-----------------|----------|----------|-----------|--------|")

    for level_name, m in results:
        lines.append(
            f"| `{level_name}` "
            f"| {m['status']} "
            f"| {m['solution_length']} "
            f"| {m['time']} "
            f"| {m['expanded']} "
            f"| {m['generated']} "
            f"| {m['memory']} |"
        )

    lines.append("")
    lines.append("### Summary")
    lines.append(f"- **Total**: {len(levels)}")
    lines.append(f"- **Solved**: {solved_count} ✅")
    lines.append(f"- **Timeout**: {timeout_count} ⏱️")
    lines.append(f"- **Error / Unsolved**: {error_count} ❌")

    report = "\n".join(lines) + "\n"

    try:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"\n{'=' * 60}")
        print(f"  Results saved to {args.output}")
        print(f"  Score: {solved_count}/{len(levels)} solved")
        print(f"{'=' * 60}")
    except Exception as e:
        print(f"ERROR writing results file: {e}")
        sys.exit(1)

    if solved_count < len(levels):
        sys.exit(1)


if __name__ == "__main__":
    main()
